"""Plantilla Excel para cargar valores de evaluación de todas las alternativas."""

from io import BytesIO

from django.db import transaction
from django.core.exceptions import ValidationError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from .evaluacion_service import (
    build_evaluacion_schema,
    export_label_for_column,
    load_valores_map,
    save_valores_bulk,
)
from .models import Alternativa, Proyecto


HEADER_FILL = PatternFill('solid', fgColor='17365D')


def build_evaluacion_template(proyecto: Proyecto) -> bytes:
    """
    Plantilla traspuesta:
      - Filas = criterios / nodos de evaluación
      - Columnas = alternativas
    Fila 2 (oculta) guarda IDs técnicos; columna A (oculta) guarda las keys.
    """
    schema = build_evaluacion_schema(proyecto)
    columnas = schema.get('columnas') or []
    alternativas = list(proyecto.alternativas.order_by('id'))
    valores_por_alt = {
        alternativa.id: load_valores_map(alternativa.id)
        for alternativa in alternativas
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Evaluacion'
    sheet.freeze_panes = 'C3'

    # Fila 1: encabezados visibles
    sheet.cell(1, 1, 'Clave')
    sheet.cell(1, 2, 'Criterio')
    for index, alternativa in enumerate(alternativas, start=3):
        sheet.cell(1, index, alternativa.nombre)

    # Fila 2: metadatos técnicos (oculta)
    sheet.cell(2, 1, '__column_key__')
    sheet.cell(2, 2, '__label__')
    for index, alternativa in enumerate(alternativas, start=3):
        sheet.cell(2, index, alternativa.id)
    sheet.row_dimensions[2].hidden = True
    sheet.column_dimensions['A'].hidden = True

    for cell in sheet[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = HEADER_FILL

    # Filas de datos: un criterio por fila
    for row_index, column in enumerate(columnas, start=3):
        key = column['key']
        sheet.cell(row_index, 1, key)
        sheet.cell(row_index, 2, export_label_for_column(column))
        for col_index, alternativa in enumerate(alternativas, start=3):
            sheet.cell(
                row_index,
                col_index,
                valores_por_alt[alternativa.id].get(key, ''),
            )

    sheet.column_dimensions['B'].width = 36
    for index in range(3, max(3, len(alternativas) + 2) + 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = 18

    instructions = workbook.create_sheet('Instrucciones')
    instructions.append(['Carga masiva de evaluación'])
    instructions['A1'].font = Font(bold=True, size=14)
    instructions.append([
        'Cada FILA es un criterio/nodo y cada COLUMNA es una alternativa. '
        'Diligencia solo las celdas de valores. No cambies la fila técnica oculta '
        'ni la columna de claves.'
    ])
    instructions.append([
        'Al importar, los valores del archivo reemplazan la evaluación actual '
        'de cada alternativa incluida.'
    ])
    instructions.column_dimensions['A'].width = 110

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _import_evaluacion_por_filas_criterio(sheet, proyecto: Proyecto) -> dict:
    """Formato actual: filas = criterios, columnas = alternativas."""
    meta = [sheet.cell(2, index).value for index in range(1, sheet.max_column + 1)]
    if meta[:2] != ['__column_key__', '__label__']:
        raise ValidationError('La plantilla fue modificada y no conserva su estructura.')

    schema = build_evaluacion_schema(proyecto)
    valid_keys = {column['key'] for column in schema.get('columnas') or []}

    alt_columns = []
    for index, raw_id in enumerate(meta[2:], start=3):
        if raw_id in (None, ''):
            continue
        try:
            alternativa = Alternativa.objects.get(pk=int(raw_id), proyecto=proyecto)
        except (Alternativa.DoesNotExist, TypeError, ValueError):
            raise ValidationError(
                f'Columna {index}: la alternativa {raw_id} no pertenece al proyecto.'
            ) from None
        alt_columns.append((index, alternativa))

    if not alt_columns:
        raise ValidationError('La plantilla no contiene alternativas válidas.')

    valores_por_alt = {alternativa.id: {} for _, alternativa in alt_columns}
    for row in range(3, sheet.max_row + 1):
        key = sheet.cell(row, 1).value
        if key not in valid_keys:
            continue
        for column_index, alternativa in alt_columns:
            raw = sheet.cell(row, column_index).value
            valores_por_alt[alternativa.id][key] = '' if raw is None else str(raw)

    updated = 0
    values_saved = 0
    for alternativa_id, valores in valores_por_alt.items():
        save_valores_bulk(alternativa_id, valores)
        updated += 1
        values_saved += sum(1 for value in valores.values() if value != '')

    return {
        'alternativas_actualizadas': updated,
        'valores_guardados': values_saved,
    }


def _import_evaluacion_legacy_por_filas_alternativa(sheet, proyecto: Proyecto) -> dict:
    """Formato anterior: filas = alternativas, columnas = criterios."""
    keys = [sheet.cell(2, index).value for index in range(1, sheet.max_column + 1)]
    if keys[:2] != ['__alternativa_id__', '__alternativa_nombre__']:
        raise ValidationError('La plantilla fue modificada y no conserva su estructura.')

    schema = build_evaluacion_schema(proyecto)
    valid_keys = {column['key'] for column in schema.get('columnas') or []}
    value_columns = [
        (index, key)
        for index, key in enumerate(keys[2:], start=3)
        if key in valid_keys
    ]
    if not value_columns:
        raise ValidationError('La plantilla no contiene columnas de evaluación válidas.')

    updated = 0
    values_saved = 0
    errors = []
    for row in range(3, sheet.max_row + 1):
        alternativa_id = sheet.cell(row, 1).value
        if alternativa_id in (None, ''):
            continue
        try:
            alternativa = Alternativa.objects.get(
                pk=int(alternativa_id),
                proyecto=proyecto,
            )
        except (Alternativa.DoesNotExist, TypeError, ValueError):
            errors.append(f'Fila {row}: la alternativa {alternativa_id} no pertenece al proyecto.')
            continue

        valores = {}
        for column_index, key in value_columns:
            raw = sheet.cell(row, column_index).value
            valores[key] = '' if raw is None else str(raw)
        save_valores_bulk(alternativa.id, valores)
        updated += 1
        values_saved += sum(1 for value in valores.values() if value != '')

    if errors:
        raise ValidationError(errors)
    return {
        'alternativas_actualizadas': updated,
        'valores_guardados': values_saved,
    }


@transaction.atomic
def import_evaluacion_template(proyecto: Proyecto, uploaded_file) -> dict:
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except Exception as exc:
        raise ValidationError('El archivo no es un Excel .xlsx válido.') from exc

    if 'Evaluacion' not in workbook.sheetnames:
        raise ValidationError('El archivo no contiene la hoja Evaluacion.')
    sheet = workbook['Evaluacion']
    marker = sheet.cell(2, 1).value

    if marker == '__column_key__':
        return _import_evaluacion_por_filas_criterio(sheet, proyecto)
    if marker == '__alternativa_id__':
        # Compatibilidad con plantillas descargadas antes de invertir ejes.
        return _import_evaluacion_legacy_por_filas_alternativa(sheet, proyecto)

    raise ValidationError(
        'La plantilla fue modificada y no conserva su estructura '
        '(falta la fila técnica de claves).'
    )
