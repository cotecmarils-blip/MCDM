from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook

from api.evaluacion_bulk_io import (
    build_evaluacion_template,
    import_evaluacion_template,
)
from api.madm_choices import simulacion_opciones_payload
from api.models import Alternativa, Proyecto


User = get_user_model()


class EvaluacionBulkIoTests(TestCase):
    def setUp(self):
        self.proyecto = Proyecto.objects.create(nombre='Proyecto Excel')
        self.alternativa = Alternativa.objects.create(
            proyecto=self.proyecto,
            nombre='Alternativa A',
        )
        self.schema = {
            'columnas': [{
                'key': 'nodo_arbol:1:1',
                'terminal_nombre': 'Velocidad',
                'escenario_nombre': 'Estándar',
            }],
        }

    @patch('api.evaluacion_bulk_io.load_valores_map')
    @patch('api.evaluacion_bulk_io.build_evaluacion_schema')
    def test_template_incluye_alternativas_y_valores(self, build_schema, load_values):
        build_schema.return_value = self.schema
        load_values.return_value = {'nodo_arbol:1:1': '12.5'}

        content = build_evaluacion_template(self.proyecto)
        sheet = load_workbook(BytesIO(content), data_only=True)['Evaluacion']

        # Filas = criterios, columnas = alternativas
        self.assertEqual(sheet.cell(1, 2).value, 'Criterio')
        self.assertEqual(sheet.cell(1, 3).value, 'Alternativa A')
        self.assertEqual(sheet.cell(2, 1).value, '__column_key__')
        self.assertEqual(sheet.cell(2, 3).value, self.alternativa.id)
        self.assertEqual(sheet.cell(3, 1).value, 'nodo_arbol:1:1')
        self.assertEqual(sheet.cell(3, 3).value, '12.5')

    @patch('api.evaluacion_bulk_io.save_valores_bulk')
    @patch('api.evaluacion_bulk_io.load_valores_map', return_value={})
    @patch('api.evaluacion_bulk_io.build_evaluacion_schema')
    def test_import_actualiza_alternativa_existente(
        self,
        build_schema,
        _load_values,
        save_values,
    ):
        build_schema.return_value = self.schema
        content = build_evaluacion_template(self.proyecto)
        workbook = load_workbook(BytesIO(content))
        # Fila de criterio, columna de alternativa
        workbook['Evaluacion'].cell(3, 3).value = 42
        uploaded = BytesIO()
        workbook.save(uploaded)
        uploaded.seek(0)

        result = import_evaluacion_template(self.proyecto, uploaded)

        self.assertEqual(result['alternativas_actualizadas'], 1)
        save_values.assert_called_once_with(
            self.alternativa.id,
            {'nodo_arbol:1:1': '42'},
        )


class AlternativasActivasTests(TestCase):
    def test_opciones_solo_selecciona_activas_por_defecto(self):
        proyecto = Proyecto.objects.create(nombre='Proyecto cálculo')
        activa = Alternativa.objects.create(proyecto=proyecto, nombre='Activa')
        Alternativa.objects.create(
            proyecto=proyecto,
            nombre='Inactiva',
            activa=False,
        )

        payload = simulacion_opciones_payload(proyecto)

        self.assertEqual(payload['defaults']['alternativa_ids'], [activa.id])
        self.assertEqual(payload['total_alternativas'], 1)
        self.assertEqual(len(payload['alternativas']), 2)
